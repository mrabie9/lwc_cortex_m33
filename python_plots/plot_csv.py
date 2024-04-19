import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import time
from natsort import natsorted
import openpyxl as xl
from enum import Enum

global x_start_e, x_stop_e, x_start_d, x_stop_d, n_loop, energy_e, energy_d
global i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated, t_calibrations
global n_lcutoff_points, l_cut_off, n_hcutoff_points, h_cut_off, calibrated
global values 

variables = ['n_loop', 'Average encryption time', 'Average decryption time', 'Average encryption energy', 'Average decryption energy',
             'Maximum encryption current', 'Average encryption current', 'Minimum encryption current', 
             'Maximum decryption current', 'Average decryption current', 'Minimum decryption current',
             'Calibrations', 'Calibration time(s)']

# values = [n_loop, t_avg_e, t_avg_d,  energy_e, energy_d,
#           i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated,', '.join(map(str, t_calibrations))]


# Contains column value for the AUT
class AUT(Enum):
    ascon128_enc_dec_1000x = 3
    ascon128Armv7_enc_dec_1000x = 4
    ascon128a_enc_dec_1000x = 5
    ascon128aArmv7_enc_dec_1000x = 6
    isapa128_enc_dec_500x = 7
    isapa128Armv7_enc_dec_500x = 8
    isapa128a_enc_dec_500x = 9
    isapa128aArmv7_enc_dec_500x = 10
    sparkle128_enc_dec_1000x = 11
    sparkle128Armv7_enc_dec_1000x = 12
    sparkle256_enc_dec_1000x = 13
    sparkle256Armv7_enc_dec_1000x = 14
    tinyjambu_enc_dec_1000x = 15
    tinyjambuOpt_enc_dec_1000x = 16
    giftc_enc_dec_200x = 17
    xoodyak_enc_dec_1000x = 18
    romulusn_enc_dec_50x = 19
    romulusnOpt_enc_dec_50x = 20
    eleph_enc_dec_10x = 21
    grain_enc_dec_10x = 22
    photon_enc_dec_15x = 23

# Start col for xlsx write
col = AUT.ascon128_enc_dec_1000x.value 

# Start/end times dict
# data_obtained = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# energy_calc = False#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/ascon128a2_enc_dec_1000x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/ascon128a2_enc_dec_1000x.txt"
# timestamps = {
#     # 'ascon128_enc_dec_1000x': {'start_e': 0.88064, 'stop_e':86.84772 , 'start_d': 89.84808, 'stop_d':174.49783},
#     # 'ascon128a_enc_dec_1000x': {'start_e': 0.52947, 'stop_e':60.11307, 'start_d': 63.11367 , 'stop_d':122.25007},
#     # 'ascon128a2_enc_dec_1000x': {'start_e': 0.71895, 'stop_e': 59.99247, 'start_d':  62.99328, 'stop_d': 122.14987},
#     # 'ascon128a3_enc_dec_1000x': {'start_e': 1.18470, 'stop_e': 59.98670, 'start_d':  62.98678, 'stop_d': 122.63684},
#     # 'giftc_enc_dec_200x': {'start_e': 0.85437, 'stop_e': 156.25025, 'start_d': 159.25074, 'stop_d': 314.85193},
#     'giftc2_enc_dec_200x': {'start_e': 1.25991, 'stop_e': 159.41488, 'start_d': 162.41518, 'stop_d': 320.73839},
#     # 'isapa128_enc_dec_500x': {'start_e': 0.94348, 'stop_e': 141.19371, 'start_d': 144.19465, 'stop_d': 284.03599},
#     # 'isapa128a_enc_dec_500x': {'start_e': 0.99050, 'stop_e': 116.92109, 'start_d':  119.92170, 'stop_d': 235.52180},
#     # 'sparkle128_enc_dec_1000x': {'start_e': 1.22652, 'stop_e': 80.71972, 'start_d': 83.72060, 'stop_d': 163.99806},
#     # 'sparkle256_enc_dec_1000x': {'start_e': 1.04642, 'stop_e': 115.44641, 'start_d': 118.44648, 'stop_d': 223.41102},
#     # 'tinyjambu_enc_dec_1000x': {'start_e': 1.06393, 'stop_e': 136.55366, 'start_d': 139.55458, 'stop_d': 274.98104},
#     # 'xoodyak_enc_dec_1000x': {'start_e': 0.76751, 'stop_e': 217.19979, 'start_d': 220.20037, 'stop_d': 437.68150},
#     # 'eleph_enc_dec_10x': {'start_e': 0.97812, 'stop_e': 120.75210, 'start_d': 123.75230, 'stop_d': 240.38971},
#     # 'grain_enc_dec_10x': {'start_e': 1.06599, 'stop_e': 108.55892, 'start_d': 110.37643, 'stop_d': 211.14548},
#     # 'photon_enc_dec_15x': {'start_e': 1.43538, 'stop_e': 176.73859, 'start_d': 179.73937, 'stop_d': 355.95881},
#     # 'romulusn_enc_dec_50x': {'start_e': 1.32967, 'stop_e': 138.96308, 'start_d': 141.96324, 'stop_d': 279.46629},
    
#     # 'grain2_enc_dec_10x': {'start_e': 1.32234, 'stop_e': 112.35972, 'start_d': 115.36006, 'stop_d': 224.39224},
#     # 'eleph2_enc_dec_10x': {'start_e': 1.18843, 'stop_e': 121.78665, 'start_d': 124.78685, 'stop_d': 245.29173},
#     # 'sparkle2562_enc_dec_1000x': {'start_e': 1.07422, 'stop_e': 115.47440, 'start_d': 118.47448, 'stop_d': 233.43923},

#     # 'eleph3_enc_dec_10x': {'start_e': 7.78885, 'stop_e': 144.36973, 'start_d': 147.37041, 'stop_d': 283.96761},

#     ### 'Random' second sets
#     # 'photon2_enc_dec_15x': {'start_e': 0.78681, 'stop_e': 202.93627, 'start_d': 205.93713, 'stop_d': 407.90221},
#     # 'ascon128a_ngnd_enc_dec_1000x': {'start_e': 0., 'stop_e': , 'start_d': , 'stop_d': }
#     }

data_obtained = True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
energy_calc = True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#FalseFalse#True#False#True#False
# data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O0/isapa128a_enc_dec_500x/"
# output = "../Data/Power/python_plots/00_Output/02_run2/O0/isapa128a_enc_dec_500x.txt"
row = 3
timestamps = { # regex : \{.*
    ## Run 2 start: O0 (with sync) 
    # 'O0': {
    #     'row' : 3,
    #     'ascon128_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'ascon128a_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'isapa128_enc_dec_500x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'isapa128a_enc_dec_500x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'sparkle128_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'sparkle256_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'tinyjambu_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'giftc_enc_dec_200x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'xoodyak_enc_dec_1000x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'romulusn_enc_dec_50x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'eleph_enc_dec_10x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'grain_enc_dec_10x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    #     'photon_enc_dec_15x': {'start_e': , 'stop_e': , 'start_d': , 'stop_d': }, 
    # },

    'O2': {
        'row' : 20,
        'ascon128_enc_dec_1000x': {'start_e': 2.37403, 'stop_e': 11.93094, 'start_d': 14.90399, 'stop_d': 24.74912}, 
        'ascon128Armv7_enc_dec_1000x': {'start_e': 2.41037, 'stop_e': 7.12427, 'start_d': 10.09707, 'stop_d': 14.78279}, 
        'ascon128a_enc_dec_1000x': {'start_e': 2.26570, 'stop_e': 9.35680, 'start_d': 13.32949, 'stop_d': 19.59281}, 
        'ascon128aArmv7_enc_dec_1000x': {'start_e': 2.46397, 'stop_e': 5.60806, 'start_d': 8.58053, 'stop_d': 11.69660}, 
        'isapa128_enc_dec_500x': {'start_e': 1.94453, 'stop_e': 35.74720, 'start_d': 38.71993, 'stop_d': 72.51813}, 
        'isapa128Armv7_enc_dec_500x': {'start_e': 2.18584, 'stop_e': 10.23172, 'start_d': 13.20501, 'stop_d': 23.23334}, 
        'isapa128a_enc_dec_500x': {'start_e': 2.53140, 'stop_e': 30.96853, 'start_d': 33.94171, 'stop_d': 62.37519}, 
        'isapa128aArmv7_enc_dec_500x': {'start_e': 2.56111, 'stop_e': 8.49112, 'start_d': 11.47124, 'stop_d': 19.39531}, 
        'sparkle128_enc_dec_1000x': {'start_e': 1.95622, 'stop_e': 9.52979, 'start_d': 12.50296, 'stop_d': 20.30800}, 
        'sparkle128Armv7_enc_dec_1000x': {'start_e': 2.58273, 'stop_e': 5.17262, 'start_d': 8.14516, 'stop_d': 10.73963}, 
        'sparkle256_enc_dec_1000x': {'start_e': 2.45116, 'stop_e': 12.88297, 'start_d': 15.85568, 'stop_d': 26.53489}, 
        'sparkle256Armv7_enc_dec_1000x': {'start_e': 2.64624, 'stop_e': 6.67907, 'start_d': 9.65174, 'stop_d': 13.64965}, 
        'tinyjambu_enc_dec_1000x': {'start_e': 2.59973, 'stop_e': 22.06532, 'start_d': 25.03722, 'stop_d': 44.44419}, 
        'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.83795, 'stop_e': 13.80457, 'start_d': 16.77720, 'stop_d': 28.68823}, 
        'giftc_enc_dec_200x': {'start_e': 2.49195, 'stop_e': 44.29604, 'start_d': 47.26853, 'stop_d': 89.12032}, 
        'xoodyak_enc_dec_1000x': {'start_e': 2.44206, 'stop_e': 41.17063, 'start_d': 44.14307, 'stop_d': 82.73165}, 
        'romulusn_enc_dec_50x': {'start_e': 2.17843, 'stop_e': 16.07612, 'start_d': 19.06999, 'stop_d': 35.09039}, #extremely sus
        'romulusnOpt_enc_dec_50x': {'start_e': 1.95808, 'stop_e': 6.12440, 'start_d': 9.21669, 'stop_d': 15.21800}, #extremely
        'eleph_enc_dec_10x': {'start_e': 2.38288, 'stop_e': 32.31490, 'start_d': 35.28806, 'stop_d': 65.21822}, 
        'grain_enc_dec_10x': {'start_e': 2.45870, 'stop_e': 30.85500, 'start_d': 33.83369, 'stop_d': 62.14700}, 
        'photon_enc_dec_15x': {'start_e': 1.91234, 'stop_e': 33.09888, 'start_d': 36.07207, 'stop_d': 67.25459}, 
    },

    'O3': {
        ## O3 (with sync)
        'row' : 37,
        'ascon128_enc_dec_1000x': {'start_e': 1.02250, 'stop_e': 8.81070, 'start_d': 11.78294, 'stop_d': 19.62652}, 
        'ascon128Armv7_enc_dec_1000x': {'start_e': 1.39305, 'stop_e': 6.10744, 'start_d': 9.08072, 'stop_d': 13.76701}, #SUS
        'ascon128a_enc_dec_1000x': {'start_e': 1.37368, 'stop_e': 6.60820, 'start_d': 9.58163, 'stop_d': 14.90931}, 
        'ascon128aArmv7_enc_dec_1000x': {'start_e': 1.25941, 'stop_e': 7.50325, 'start_d': 10.49183, 'stop_d': 15.50772}, #SUS
        'isapa128_enc_dec_500x': {'start_e': 1.29696, 'stop_e': 25.53046, 'start_d': 28.50220, 'stop_d': 52.70238}, 
        'isapa128Armv7_enc_dec_500x': {'start_e': 1.31951, 'stop_e': 9.32562, 'start_d': 12.29804, 'stop_d': 20.30366}, 
        'isapa128a_enc_dec_500x': {'start_e': 0.76078, 'stop_e': 19.73634, 'start_d': 22.70885, 'stop_d': 41.68275}, 
        'isapa128aArmv7_enc_dec_500x': {'start_e': 0.27230, 'stop_e': 5.70411, 'start_d': 8.67663, 'stop_d': 14.57444}, #SUS start
        'sparkle128_enc_dec_1000x': {'start_e': 0.80780, 'stop_e': 7.56849, 'start_d': 10.54068, 'stop_d': 17.74982}, #SUS start
        'sparkle128Armv7_enc_dec_1000x': {'start_e': 1.41833, 'stop_e': 3.90840, 'start_d': 6.88102, 'stop_d': 9.37518}, 
        'sparkle256_enc_dec_1000x': {'start_e': 1.57660, 'stop_e': 11.24437, 'start_d': 14.21624, 'stop_d': 24.08220}, 
        'sparkle256Armv7_enc_dec_1000x': {'start_e': 1.50846, 'stop_e': 5.54791, 'start_d': 8.52011, 'stop_d': 12.54082}, 
        'tinyjambu_enc_dec_1000x': {'start_e': 1.41202, 'stop_e': 17.25013, 'start_d': 20.22312, 'stop_d': 36.05926}, 
        'tinyjambuOpt_enc_dec_1000x': {'start_e': 1.39317, 'stop_e': 13.33886, 'start_d': 16.31070, 'stop_d': 28.23758}, 
        'giftc_enc_dec_200x': {'start_e': 1.29737, 'stop_e': 20.09049, 'start_d': 23.06211, 'stop_d': 41.84347}, 
        'xoodyak_enc_dec_1000x': {'start_e': 1.48747, 'stop_e': 15.61579, 'start_d': 18.58829, 'stop_d': 32.56706}, 
        'romulusn_enc_dec_50x': {'start_e': 1.36210, 'stop_e': 6.62992, 'start_d': 9.60220, 'stop_d': 14.87715}, 
        'romulusnOpt_enc_dec_50x': {'start_e': 1.01335, 'stop_e': 2.16433, 'start_d': 5.13682, 'stop_d': 6.29749}, 
        'eleph_enc_dec_10x': {'start_e': 0.75621, 'stop_e': 20.14343, 'start_d': 23.11606, 'stop_d': 42.50315}, 
        'grain_enc_dec_10x': {'start_e': 0.88573, 'stop_e': 28.13817, 'start_d': 31.10991, 'stop_d': 58.24122}, 
        'photon_enc_dec_15x': {'start_e': 1.51557, 'stop_e': 19.81075, 'start_d': 22.77781, 'stop_d': 41.06898}, 

    }, 

    'Os': {
    #     ## Os (with sync)
        'row' : 54,
        'ascon128_enc_dec_1000x': {'start_e': 2.41304, 'stop_e': 15.75085, 'start_d': 18.72286, 'stop_d': 32.17177}, 
        'ascon128Armv7_enc_dec_1000x': {'start_e': 1.86693, 'stop_e': 6.56237, 'start_d': 9.53527, 'stop_d': 14.28603}, 
        'ascon128a_enc_dec_1000x': {'start_e': 2.56514, 'stop_e': 12.48843, 'start_d': 15.46047, 'stop_d': 25.48190}, 
        'ascon128aArmv7_enc_dec_1000x': {'start_e': 2.45464, 'stop_e': 5.54786, 'start_d': 8.52029, 'stop_d': 11.61796},
        'isapa128_enc_dec_500x': {'start_e': 2.52457, 'stop_e': 43.73854, 'start_d': 46.71072, 'stop_d': 87.92560}, 
        'isapa128Armv7_enc_dec_500x': {'start_e': 2.48072, 'stop_e': 10.64315, 'start_d': 13.61578, 'stop_d': 21.77845}, 
        'isapa128a_enc_dec_500x': {'start_e': 2.33187, 'stop_e': 37.52490, 'start_d': 40.49805, 'stop_d': 75.68812}, 
        'isapa128aArmv7_enc_dec_500x': {'start_e': 2.51144, 'stop_e': 8.55285, 'start_d': 11.52593, 'stop_d': 17.56767}, 
        'sparkle128_enc_dec_1000x': {'start_e': 2.35007, 'stop_e': 10.37047, 'start_d': 13.34325, 'stop_d': 21.65767}, # SUS END 
        'sparkle128Armv7_enc_dec_1000x': {'start_e': 2.52977, 'stop_e': 5.11364, 'start_d': 8.08648, 'stop_d': 10.67053}, 
        'sparkle256_enc_dec_1000x': {'start_e': 2.07569, 'stop_e': 13.29187, 'start_d': 16.26426, 'stop_d': 27.77032}, 
        'sparkle256Armv7_enc_dec_1000x': {'start_e': 2.37905, 'stop_e': 6.39388, 'start_d': 9.26414, 'stop_d': 15.24520}, 
        'tinyjambu_enc_dec_1000x': {'start_e': 2.46432, 'stop_e': 24.90874, 'start_d': 27.88129, 'stop_d': 50.30585}, 
        'tinyjambuOpt_enc_dec_1000x': {'start_e': 2.55719, 'stop_e': 16.51569, 'start_d': 19.48920, 'stop_d': 33.42918}, 
        'giftc_enc_dec_200x': {'start_e': 1.97305, 'stop_e': 41.96765, 'start_d': 44.94041, 'stop_d': 84.93631}, 
        'xoodyak_enc_dec_1000x': {'start_e': 2.59438, 'stop_e': 43.55850, 'start_d': 46.53142, 'stop_d': 87.35707}, 
        'romulusn_enc_dec_50x': {'start_e': 2.50744, 'stop_e': 19.61178, 'start_d': 22.58508, 'stop_d': 39.70875}, 
        'romulusnOpt_enc_dec_50x': {'start_e': 2.51312, 'stop_e': 3.66764, 'start_d': 6.64118, 'stop_d': 7.79345}, 
        'eleph_enc_dec_10x': {'start_e': 2.44766, 'stop_e': 39.11380, 'start_d': 42.08640, 'stop_d': 78.74528}, 
        'grain_enc_dec_10x': {'start_e': 2.46674, 'stop_e': 33.17888, 'start_d': 36.15913, 'stop_d': 66.77305}, 
        'photon_enc_dec_15x': {'start_e': 2.55105, 'stop_e': 37.63255, 'start_d': 40.60511, 'stop_d': 75.68681}, 
    }    
}

plot_all_bool = False
# apps = ["ascon128_enc_dec_1000x", "ascon128Armv7_enc_dec_1000x", "ascon128a_enc_dec_1000x", "ascon128aArmv7_enc_dec_1000x"]
# apps = ["isapa128_enc_dec_500x", "isapa128Armv7_enc_dec_500x", "isapa128a_enc_dec_500x", "isapa128aArmv7_enc_dec_500x", ]
# apps = [ "sparkle128_enc_dec_1000x", "sparkle128Armv7_enc_dec_1000x", "sparkle256_enc_dec_1000x", "sparkle256Armv7_enc_dec_1000x"]
# apps = ["tinyjambu_enc_dec_1000x", "tinyjambuOpt_enc_dec_1000x", "giftc_enc_dec_200x", "xoodyak_enc_dec_1000x",]
apps = [ "romulusn_enc_dec_50x" , "romulusnOpt_enc_dec_50x", "eleph_enc_dec_10x","grain_enc_dec_10x", "photon_enc_dec_15x"]

# apps  = ["ascon128Armv7_enc_dec_1000x"]
def plot_all():
    dfs = []
    t = time.time()
    for app in apps:
        print("Reading ", app, apps.index(app))
        data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/O2/" + app + "/"
        files = os.listdir(data_dir)

        # Filter out only CSV files
        csv_files = [file for file in files if (file.endswith('.csv') and not "summary" in file)]
        csv_files = natsorted(csv_files)

        
        for file_name in csv_files:
            file_path = os.path.join(data_dir, file_name)
            # Read the CSV file into a DataFrame
            df = pd.read_csv(file_path, delimiter=';', names=[app + "x", app + "y"])
            dfs.append(df)

        # Concatenate all DataFrames into a single DataFrame ignore index HAS to be true
        df = pd.concat(dfs, ignore_index=True)
        # x_values[apps.index(app)] = df['x']/1e3
        # y_values[apps.index(app)] = df['y']/1e6
    print(time.time() - t)
    for app in apps:
        t = time.time()
        x_values= df[app + 'x']/1e3
        y_values = df[app + 'y']/1e6
        # Plotting the graph
        plt.plot(x_values, y_values, linestyle='-') 
        # plt.figure()
        plt.xlabel('Time (s)') 
        plt.ylabel('Current (A)') 
        plt.title(app) 
        # plt.text(4.28073, 0.047, "Start of program", fontsize=12, ha='center', va='center', color='black')
        plt.grid(True)
        print(time.time() - t)
        plt.show()

def main(ts_dict):
    global x_start_e, x_stop_e, x_start_d, x_stop_d, n_loop, energy_e, energy_d
    global i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated, t_calibrations
    global n_lcutoff_points, l_cut_off, n_hcutoff_points, h_cut_off, calibrated
    global values
    # List all files in the directory
    files = os.listdir(data_dir)

    # Filter out only CSV files
    csv_files = [file for file in files if (file.endswith('.csv') and not "summary" in file)]
    csv_files = natsorted(csv_files)
    # print(csv_files)

    # Loop through each CSV file and read its contents
    dfs = []
    for file_name in csv_files:
        file_path = os.path.join(data_dir, file_name)
        # Read the CSV file into a DataFrame
        df = pd.read_csv(file_path, delimiter=';', names=["x", "y"])
        dfs.append(df)

    # Concatenate all DataFrames into a single DataFrame ignore index HAS to be true
    df = pd.concat(dfs, ignore_index=True)
    # Assuming your CSV file has columns named 'x' and 'y', change them accordingly if they are different
    x_values = df['x']/1e3
    y_values = df['y']/1e6
    # print(x_values)

    # Check for calibrations
    n_calibrated = len(y_values[y_values>1])
    t_calibrations = y_values[y_values>1].index
    calibrated = bool(n_calibrated)
    # print("Number of calibrations: ", n_calibrated)

    # remove calibration current samples
    l_cut_off = []
    h_cut_off = 1
    n_hcutoff_points =  len(y_values[y_values>h_cut_off])
    # print("Calibrated: ", calibrated)
    if len(y_values[y_values<0.005]) < 10:
        l_cut_off = 0.005
        n_lcutoff_points =  len(y_values[y_values<l_cut_off])
        x_values = x_values[y_values<1]
        x_values = x_values[y_values>0.005]
        y_values = y_values[y_values<1]
        y_values = y_values[y_values>0.005]
        # print("Removed " + str(n_lcutoff_points) + " at 0.005 A cut-off")
        # print("Removed " + str(n_hcutoff_points) + " above " + str(n_hcutoff_points) + "A")
    else:
        l_cut_off = 0.001
        n_lcutoff_points =  len(y_values[y_values<l_cut_off])
        x_values = x_values[y_values<1]
        x_values = x_values[y_values>0.001]
        y_values = y_values[y_values<1]
        y_values = y_values[y_values>0.001]
        # print("Removed " + str(n_lcutoff_points) + " below " + str(l_cut_off) + "A")
        # print("Removed " + str(n_hcutoff_points) + " above " + str(h_cut_off) + "A")

    ## get start/end time - for report
    # specific_x_val = 3.99519
    # index_of_x = x_values[x_values== specific_x_val].index[0]
    # x_values = x_values[index_of_x:]
    # y_values = y_values[index_of_x:]
    # annotate_x = x_values[index_of_x]
    # annotate_y = round(y_values[index_of_x], 4)
    # # Annotate the point with its coordinates
    # plt.annotate(f'({annotate_x}, {annotate_y})', xy=(annotate_x, annotate_y), xytext=(annotate_x-2e-5, annotate_y+1e-4))
    # plt.xlim(x_values[index_of_x]-9e-5,x_values[index_of_x]+5e-5 )
    # plt.gca().yaxis.set_ticks_position('right')
    # plt.gca().yaxis.set_label_position('right')

    # Figure out n_loop
    if "1000" in data_dir:
        n_loop = 1000
    elif "500" in data_dir:
        n_loop = 500
    elif "200" in data_dir:
        n_loop = 200
    elif "100" in data_dir:
        n_loop = 100
    elif "50" in data_dir:
        n_loop = 50
    elif "15" in data_dir:
        n_loop = 15
    elif "10" in data_dir:
        n_loop = 10
    else: 
        print("N_loop not recognised")
        os._exit

    if energy_calc:
        ## try nested dict
        x_start_e = 0
        x_stop_e = 0
        x_start_d = 0
        x_stop_d = 0

        for app, data in ts_dict.items():
            if app in data_dir:
                x_start_e = data['start_e']
                x_stop_e = data['stop_e']
                x_start_d = data['start_d']
                x_stop_d = data['stop_d']

        # energy calculation - encryption
        # x_start_e = 1.180800
        # x_stop_e = 182.506460
        id_x_start_e = x_values[round(x_values,5)== x_start_e].index[0]
        # print("Start enc time: ", x_values[id_x_start_e])
        id_x_stop_e = x_values[round(x_values,5)== x_stop_e].index[0]
        i_e_avg = round(y_values[id_x_start_e:id_x_stop_e].mean(),5)
        i_e_max = round(y_values[id_x_start_e:id_x_stop_e].max(),5)
        i_e_min = round(y_values[id_x_start_e:id_x_stop_e].min(),5)
        energy_e = i_e_avg * 3.3 * (x_stop_e-x_start_e)/n_loop
        energy_e = round(energy_e,5)
        # print("Average energy - encryption: ", energy_e, (x_stop_e-x_start_e)/n_loop)

        # energy calculation - decryption
        # x_start_d = 185.512820
        # x_stop_d = 367.872190
        id_x_start_d = x_values[round(x_values,5)== x_start_d].index[0]
        id_x_stop_d = x_values[round(x_values,5)== x_stop_d].index[0]
        # print("Start dec time: ", x_values[id_x_start_d])
        i_d_avg = round(y_values[id_x_start_d:id_x_stop_d].mean(),5)
        i_d_max = round(y_values[id_x_start_d:id_x_stop_d].max(),5)
        i_d_min = round(y_values[id_x_start_d:id_x_stop_d].min(),5)
        energy_d = i_d_avg * 3.3 * (x_stop_d-x_start_d)/n_loop
        energy_d = round(energy_d,5)
        # print("Average energy - decryption: ", energy_d, x_start_d, id_x_start_d, x_stop_d, id_x_stop_d, (x_stop_d-x_start_d)/n_loop)

        # # store relevant values
        t_avg_e = round((x_stop_e-x_start_e)/n_loop,5)
        t_avg_d = round((x_stop_d-x_start_d)/n_loop,5)

        # Store results in a text file
        f = open(output, "w") 
        f.close()
        f = open(output, "a")
        if f.closed:
            print("File not open!!")
        f.write(output)
        f.write(": \n")
        f.write("Execution time: \n")
        f.write("\tEncryption start time:\t\t%f s\n" % x_start_e)
        f.write("\tEncryption end time:\t\t%f s\n" % x_stop_e)
        f.write("\tDecryption start time:\t\t%f s\n" % x_start_d)
        f.write("\tDecryption end time:\t\t%f s\n" % x_stop_d)
        f.write("\tAverage encryption time:\t%f s\n" % (t_avg_e))
        f.write("\tAverage decryption time:\t%f s\n\n" % (t_avg_d))
        f.write("Energy consumption: \n")
        f.write("\tAverage encryption energy: \t%f J\n" % energy_e)
        f.write("\tAverage decryption energy: \t%f J\n\n" % energy_d)
        f.write("Encryption current: \n")
        f.write("\tMaximum encryption current: %f A\n" % i_e_max)
        f.write("\tAverage encryption current: %f A\n" % i_e_avg)
        f.write("\tMinimum encryption current: %f A\n\n" % i_e_min)
        f.write("Decryption current: \n")
        f.write("\tMaximum decryption current: %f A\n" % i_d_max)
        f.write("\tAverage decryption current: %f A\n" % i_d_avg)
        f.write("\tMinimum decryption current: %f A\n\n" % i_d_min)
        f.write("Notes: \n")
        f.write("\tN_LOOP:\t\t\t\t\t\t%d\n" % n_loop)
        # f.write("\tCalibrated: \t\t\t\t%s\n" % calibrated)
        f.write("\tNum Calibrations: \t\t\t%d\n" % n_calibrated)
        f.write("\tCalibration time(s): \t\t")
        if not calibrated:
            f.write("N/A")
        else:
            for t in t_calibrations:
                f.write("%.5f s\t" % (round(t/100000,5)))
            f.write("\n\tRemoved " + str(n_lcutoff_points) + " samples(s) below " + str(l_cut_off) + "A")
            f.write(" and " + str(n_hcutoff_points) + " above " + str(h_cut_off) + "A")
        f.close
        values = [n_loop, t_avg_e, t_avg_d,  energy_e, energy_d,
          i_e_max, i_e_avg, i_e_min, i_d_max, i_d_avg, i_d_min, n_calibrated,', '.join(map(str, t_calibrations/100000))]


    else: 
        # Plotting the graph
        plt.plot(x_values, y_values, linestyle='-')  # You can customize the marker and linestyle as needed
        plt.xlabel('Time (s)')  # Replace 'X Axis Label' with your desired label
        plt.ylabel('Current (A)')  # Replace 'Y Axis Label' with your desired label
        plt.title(data_dir[-15:])  # Replace 'Title of the Graph' with your desired title


        # plt.text(4.28073, 0.047, "Start of program", fontsize=12, ha='center', va='center', color='black')
        plt.grid(True)  # Add gridlines if needed
        plt.show()


def writexl(col, row):
    global values

    # open workbook
    workbook = xl.load_workbook("../Data/Data_m33.xlsx")
    ws = workbook['Power - M33']

    # Write the variable names to the first row
    # for row, var in enumerate(variables, start=row):
    #     ws.cell(row=row, column=col-1, value=var)
    # row = row_start
    for row, var in enumerate(values, start=row):
        ws.cell(row=row, column=col, value=var)
    workbook.save("../Data/Data_m33.xlsx")

if plot_all_bool:
    plot_all()
else:
    if data_obtained:
        t = time.time()
        for opt, data in timestamps.items():
            print(opt)
            col = 3
            for app, values in data.items():
                if not 'row' in app:
                    print(app)
                    data_dir = "../Data/Power/python_plots/02_csv_dir/02_run2/" + opt +"/" + app + "/"
                    output = "../Data/Power/python_plots/00_Output/02_run2/" + opt +"/"  + app + ".txt"
                    # print(app)
                    main(data)
                    writexl(col, row)
                    col +=1
                else:
                    row = values
                    # print("Row ", values)
        print("Took: ", time.time() - t)
    else: 
        main()
